import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series


# ======================
# Experimental constants
# ======================

area = 0.65973
pH = 14
reference = "HgHgO"

if reference == "HgHgO":
    E0 = 0.098
else:
    E0 = 0.197


# ======================
# Read EC-Lab file
# ======================

def read_mpt(filepath):

    with open(filepath,"r",encoding="latin1") as f:
        lines = f.readlines()

    header=None
    for i,line in enumerate(lines):
        if "Ewe/V" in line and "<I>/mA" in line:
            header=i
            break

    headers=lines[header].strip().split("\t")

    df=pd.read_csv(
        filepath,
        sep="\t",
        names=headers,
        skiprows=header+1,
        encoding="latin1",
        engine="python"
    )

    potential=pd.to_numeric(df["Ewe/V"],errors="coerce")
    current=pd.to_numeric(df["<I>/mA"],errors="coerce")

    data=pd.DataFrame()

    data["E_RHE"]=potential + E0 + 0.059*pH
    data["j_mA_cm2"]=current/area
    data["j_A_cm2"]=data["j_mA_cm2"]/1000
    data["eta"]=-data["E_RHE"]

    data=data.dropna()

    return data


# ======================
# Compute parameters
# ======================

def compute_parameters(data):

    onset = data.loc[data["j_mA_cm2"].abs()>1].iloc[0]["E_RHE"]

    pot5 = data.iloc[(data["j_mA_cm2"].abs()-5).abs().argsort()[:1]]["E_RHE"].values[0]

    logj = np.log10(np.abs(data["j_A_cm2"]))
    eta = data["eta"]

    mask=(logj>-6)&(logj<-2)

    x=logj[mask]
    y=eta[mask]

    slope,intercept=np.polyfit(x,y,1)

    tafel_slope=slope*1000

    log_i0=-intercept/slope
    i0=10**log_i0*1000

    return onset,pot5,tafel_slope,i0


# ======================
# File paths
# ======================

files={
"bare_ss":"F/HER studies/bare SS_C02.mpt",
"mod_ss7":"F/HER studies/mSS 7 min HER after condn_C02.mpt",
"pt_acid":"F/pH study/Pt in acid_C02.mpt",
"pt_base":"F/pH study/Pt in base_C02.mpt"
}


data={k:read_mpt(v) for k,v in files.items()}

params={k:compute_parameters(v) for k,v in data.items()}


# ======================
# Create Excel workbook
# ======================

wb=Workbook()
wb.remove(wb.active)


# ======================
# Data + graph sheets
# ======================

for name,df in data.items():

    ws=wb.create_sheet(name)

    ws.append(["E_RHE (V)","j (mA/cm2)"])

    for i in range(len(df)):
        ws.append([df.iloc[i]["E_RHE"],df.iloc[i]["j_mA_cm2"]])

    chart=ScatterChart()

    chart.x_axis.title="E (V vs RHE)"
    chart.y_axis.title="j (mA cm^-2)"

    xvalues=Reference(ws,min_col=1,min_row=2,max_row=len(df)+1)
    yvalues=Reference(ws,min_col=2,min_row=2,max_row=len(df)+1)

    series=Series(yvalues,xvalues,title=name)

    chart.series.append(series)

    ws.add_chart(chart,"E2")


# ======================
# Summary sheet
# ======================

ws=wb.create_sheet("Summary")


ws.append(["Parameter","Electrode-1","Electrode-2","Electrode-3","Electrode-4"])

ws.append([
"Onset Potential",
params["bare_ss"][0],
params["mod_ss7"][0],
params["pt_base"][0],
params["pt_acid"][0]
])

ws.append([
"Potential at 5 mA cm-2",
params["bare_ss"][1],
params["mod_ss7"][1],
params["pt_base"][1],
params["pt_acid"][1]
])

ws.append([
"Tafel slope b (mV dec-1)",
params["bare_ss"][2],
params["mod_ss7"][2],
params["pt_base"][2],
params["pt_acid"][2]
])

ws.append([
"Exchange current density i0 (mA cm-2)",
params["bare_ss"][3],
params["mod_ss7"][3],
params["pt_base"][3],
params["pt_acid"][3]
])


# ======================
# Module 2 table
# ======================

ws2=wb.create_sheet("Module2")

ws2.append(["Parameter","0.5 M H2SO4","0.1 M KOH"])

ws2.append(["Onset Potential",params["pt_acid"][0],params["pt_base"][0]])

ws2.append(["Potential at 5 mA cm-2",params["pt_acid"][1],params["pt_base"][1]])

ws2.append(["Tafel slope b (mV dec-1)",params["pt_acid"][2],params["pt_base"][2]])

ws2.append(["Exchange current density i0 (mA cm-2)",params["pt_acid"][3],params["pt_base"][3]])


wb.save("HER_LSV_Final.xlsx")

print("Excel file created: HER_LSV_Final.xlsx")